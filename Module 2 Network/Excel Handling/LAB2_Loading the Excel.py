from openpyxl import *
import os,pandas


filename="Sample.xlsx"
wb=load_workbook(filename)
ws=wb.active

#print(ws['A1'].value + "\t"+ ws['B1'].value )

ws['A2']="192.168.1.221"

wb.save(filename)
