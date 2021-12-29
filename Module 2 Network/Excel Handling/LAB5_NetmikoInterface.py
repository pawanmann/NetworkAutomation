import os
import time
from netmiko import ConnectHandler
from openpyxl import Workbook,load_workbook

try:
    if "Config.xlsx" in os.listdir():
        print("#"*20 + "\nFetching data from file\n"+"#"*20)
        wb=load_workbook("Config.xlsx")
    else:
        print("#" * 20 + "\nCreating New Excel File\n" + "#" * 20)
        wb= Workbook()

    ws=wb.active

    start= time.time()
    device_list=[]
    ws['A1']='IP ADD'
    ws['B1']='Hostname'
    ws['C1']='Version'
    ws['D1']='Config'

    with open("../Threading/devices.txt") as file:
        devices=file.read().splitlines()

    for device in devices:

        User_Input={
            'device_type':'cisco_ios',
            'ip': device,
            'username':'dhiraj',
            'password':'cisco',
            'secret':'cisco'
        }
        Connection = ConnectHandler(**User_Input)
        if '>' in Connection.find_prompt():
            print("Entering into enable mode")
            Connection.enable()

        version = Connection.send_command('show version | i uptime')
        Interface= Connection.send_command('show ip int br')
        run=Connection.send_command('show run')
        prompt=Connection.find_prompt()
        Max=ws.max_row+1
        ws['A'+str(Max)]=device
        ws['B' + str(Max)] = prompt[0:-1]
        ws['C' + str(Max)] = version
        ws['D' + str(Max)] = run
        #ws['E' + str(Max)] = run

        wb.save("Config.xlsx")
        print(f"Closing Connections for {device} ")
        Connection.disconnect()


    print(f"{'#'*30}\n\tAutomation Completed\n{'#'*30}")

    end=time.time()

    print(f"Time required: {end-start}")
    os.startfile("Config.xlsx")

except PermissionError:
    print("#"*20 + "\nPlease Close the Excel\n"+"#"*20)
