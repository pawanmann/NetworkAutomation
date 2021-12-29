import time,threading,os,sys
from netmiko import ConnectHandler
from openpyxl import Workbook,load_workbook

start= time.time()

def connection(User_Input):
    Connection = ConnectHandler(**User_Input)
    if '>' in Connection.find_prompt():
        print("Entering into enable mode")
        Connection.enable()
    Excel(Connection,User_Input)

def Excel(Connection,User_Input):
    try:
        version = Connection.send_command('show version | i uptime')
        run = Connection.send_command('show run')
        prompt = Connection.find_prompt()
        Max = ws.max_row + 1
        ws['A' + str(Max)] = User_Input['ip']
        ws['B' + str(Max)] = prompt[0:-1]
        ws['C' + str(Max)] = version
        ws['D' + str(Max)] = run

        Connection.disconnect()

        wb.save("Config.xlsx")
        print(f"Closing Connections for {User_Input['ip']} ")
    except PermissionError:
        print("#" * 20 + "\nPlease Close the Excel\n" + "#" * 20)
        sys.exit()

def Threading():

    threads=[]

    with open("devices.txt") as file:
        devices = file.read().splitlines()

    for device in devices:
        User_Input={
            'device_type':'cisco_ios',
            'ip': device,
            'username':'dhiraj',
            'password':'cisco',
            'secret':'cisco'
        }
        th = threading.Thread(target=connection, args=(User_Input,))
        threads.append(th)

    for th in threads:
        th.start()
        time.sleep(0.5)
    for th in threads:
        th.join()

def Create():
    global wb,ws
    if "Config.xlsx" in os.listdir():
        print("#"*20 + "\nFetching data from file\n"+"#"*20)
        wb=load_workbook("Config.xlsx")
        ws = wb.active
    else:
        print("#" * 20 + "\nCreating New Excel File\n" + "#" * 20)
        wb= Workbook()
        ws = wb.active
        ws['A1']='IP ADD'
        ws['B1']='Hostname'
        ws['C1']='Version'
        ws['D1']='Config'
        wb.save("Config.xlsx")
    Threading()

def Complete():
    print(f"{'#' * 30}\n\tAutomation Completed\n{'#' * 30}")
    end = time.time()
    print(f"Time required: {end - start}")
    #os.startfile("Config.xlsx")

Create()
Complete()

