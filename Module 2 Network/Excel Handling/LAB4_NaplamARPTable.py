import json,os,pandas
from openpyxl import Workbook,load_workbook
from napalm import get_network_driver

def Connect(IP,User,Passwd):
    global cisco_device
    driver = get_network_driver('ios')
    optional_args = {'secret': 'cisco'}
    driver = get_network_driver('ios')
    ios = driver(hostname=IP, username=User, password=Passwd, optional_args=optional_args)
    ios.open()
    return ios

def Excel(file,ios,output):
    try:
        global wb,ws
        if file in os.listdir():
            print("Refreshing Excel....")
            wb = load_workbook(file)
            ws = wb.active
        else:
            print("#" * 20 + "\nCreating New Excel File\n" + "#" * 20)
            wb = Workbook()
            ws = wb.active
            ws['A1'] = 'Hostname'
            ws['B1']='Interface'
            ws['C1']='Mac Address'
            ws['D1']='IP Address'
            ws['E1']='Age'
        Max = ws.max_row + 1
        ws['A' + str(Max)] = cisco_device['IP']
        for item in output:
            Max = ws.max_row + 1
            ws['B' + str(Max)] = item['interface']
            ws['C' + str(Max)] = item['mac']
            ws['D' + str(Max)] = item['ip']
            ws['E' + str(Max)] = item['age']
        wb.save(file)
        ios.close()

    except PermissionError:
        print("#" * 20 + "\nPlease Close the Excel\n" + "#" * 20)
        exit(4)

def Output(ios):
        output=ios.get_arp_table()
        return output

def read():
    Read=pandas.read_excel(File)
    #df = pandas.DataFrame(Read, columns=['IP Address', 'Mac Address']).fillna(" ")
    print(Read.fillna(" "))

with open("../Threading/devices.txt") as file:
    File = "Config.xlsx"
    for line in file:
        Strip = line.strip()
        cisco_device= {'IP': Strip,'User': 'dhiraj', 'Passwd': 'cisco'}
        ios = Connect(**cisco_device)
        output=Output(ios)
        Create=Excel(File,ios,output)
        print(f"Adding data for: {Strip} ")

    if input("\nDo you want to open file[y/n]: ").lower()=='y':
        os.startfile(File)
    else:
        print("Thank You")
        print(ws['A2'].value)
        read()