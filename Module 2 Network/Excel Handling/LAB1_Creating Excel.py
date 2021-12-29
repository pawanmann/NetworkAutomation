from openpyxl import *
import os

wb=Workbook()
ws=wb.active

if "Sample.xlsx" in os.listdir():
    os.remove("Sample.xlsx")
    wb.save("Sample.xlsx")

else:
    wb.save("Sample.xlsx")

ws['A1']="IP address"
ws['B1']="Hostname"

wb.save("Sample.xlsx")